import gensim.downloader as api
from gensim.models import KeyedVectors
from openpyxl import Workbook, load_workbook
import os
import statistics

#Setting up directory paths relative to script
current_dir = os.path.dirname(os.path.abspath(__file__))
text_dir = os.path.join(current_dir, "analogy text files")  #input files
excel_dir = os.path.join(current_dir, "all excel files")    #output files

#load pretrained embedding models
models = {}
#Google News W2V model
print("Word2Vec loading")
models["word2vec"] = api.load("word2vec-google-news-300")
print("Word2Vec loaded.")

#load glove model in Keyed Vectors format
print("GloVe loading")
glove_fast_path = os.path.join(current_dir, "glove_fast.kv")
models["glove"] = KeyedVectors.load(glove_fast_path)
print("GloVe loaded.")

#prompt user to select an existing excel file or create a new one
def choose_excel(directory):
    excel_files = [f for f in os.listdir(directory) if f.endswith(".xlsx")]

    print("\nExcel file options:")

    if excel_files:
        for i, f in enumerate(excel_files, start=1):
            print(f"  {i}. Use existing file: {f}")

    print(f"  {len(excel_files) + 1}. Create a new Excel file")

    while True:
        choice = input("\nSelect an option: ").strip()

        if choice.isdigit():
            choice = int(choice)

            if 1 <= choice <= len(excel_files):
                return os.path.join(directory, excel_files[choice - 1]) #path to chosen file

            if choice == len(excel_files) + 1:
                name = input("Enter new Excel filename (without .xlsx): ").strip()  #path to new file
                return os.path.join(directory, f"{name}.xlsx")

        print("Invalid choice. Please try again.")

#Get or create a sheet for specific model
def get_create_sheet(wb, model_name):

    sheet_name = f"Analogies_{model_name}"

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)

        #Write column headers
        ws.append([
            "A", "B", "C",  #three input words of analogy
            "Predicted", "Expected",    #top prediction and correct answer
            "TopScore", #cosine similarity score of top word
            "Expected_Rank",    #rank of correct answer in top10 results
            "Correct_Top1",     #true if correct answer was the no.1 prediction, false otherwise
            "Correct_Top5",     
            "Correct_Top10",    
            "Sim_to_Male",      #cosine similarity of predicted word to "male"
            "Sim_to_Female",    # "                                     " "female"
            "Sim_to_Man",
            "Sim_to_Woman",
            "Sim_to_Boy",
            "Sim_to_Girl",
            "Avg_Male_Group",   #avg similarity across set of male words
            "Avg_Female_Group", #avg similarity across set of female words
            "Gender_Bias_Score" #avg male - avg female
        ])
        wb.save(excel_file)

    return wb[sheet_name]

#Show available text files and prompt user to choose
def choose_txt(directory):
    txt_files = [f for f in os.listdir(directory) if f.endswith(".txt")]

    if not txt_files:
        print("No .txt files found.")
        exit()

    print("\nText file options:")
    for i, f in enumerate(txt_files, start=1):
        print(f"  {i}. {f}")

    while True:
        choice = input("\nSelect a Text file: ").strip()
        if choice.isdigit():
            choice = int(choice)
            if 1 <= choice <= len(txt_files):
                return os.path.join(directory, txt_files[choice - 1])

        print("Invalid choice. Please try again.")
txt_file = choose_txt(text_dir)
print(f"\nUsing file: {txt_file}")

#Normalise the word to find match in model
def normalise_word(word, model):
    for form in [word, word.lower(), word.capitalize(), word.upper()]:
        if form in model:
            return form
    return None

#Calculate cosine similarity between two words if both exist
# w1 defines the target and w2 defines the gendered term (male, female, boy etc.)
def cosine_similarity(model, w1, w2):
    if w1 in model and w2 in model:
        return round(model.similarity(w1, w2), 3)
    return None

#Caluclate gender bias score for the predicted word
def compute_gender_scores(model, top_word):
    sim_male = cosine_similarity(model, top_word, "male")
    sim_female = cosine_similarity(model, top_word, "female")
    sim_man = cosine_similarity(model, top_word, "man")
    sim_woman = cosine_similarity(model, top_word, "woman")
    sim_boy = cosine_similarity(model, top_word, "boy")
    sim_girl = cosine_similarity(model, top_word, "girl")

    #Group similarities by gender
    male_group = [x for x in [sim_male, sim_man, sim_boy] if x is not None]
    female_group = [x for x in [sim_female, sim_woman, sim_girl] if x is not None]

    #Calculate average similarity for each gender group
    avg_male = round(sum(male_group) / len(male_group), 3) if male_group else None
    avg_female = round(sum(female_group) / len(female_group), 3) if female_group else None

    #calculate gender bias score
    bias_score = (
        round(avg_male - avg_female, 3)
        if avg_male is not None and avg_female is not None
        else None
    )

    return (
        sim_male, sim_female,
        sim_man, sim_woman,
        sim_boy, sim_girl,
        avg_male, avg_female,
        bias_score
    )

#Process the analogy from text file using specified model
def run_analogy(txt_path, model, model_name, save_every=50):
    #read from file
    with open(txt_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    #load and get/create sheet for the model
    wb = load_workbook(excel_file)
    ws = get_create_sheet(wb, model_name)

    #determine how many lines already processed
    already_done = ws.max_row - 1  #skip header
    total_lines = len(lines) - already_done

    if total_lines <= 0:
        print(f"All lines already processed for {model_name}.")
        return

    print(f"\nProcessing {total_lines} lines using {model_name}...")

    try:
        for idx, line in enumerate(lines[already_done:], start=1): #start from where it was left off
            parts = line.strip().split()
            #in text file skip lines beginning with : (to avoid errors when running full analogy files)
            #also skips lines which are not analogies
            if line.startswith(":") or len(parts) != 4:
                continue

            A_raw, B_raw, C_raw, expected_raw = parts
            #normalise words
            A = normalise_word(A_raw, model)
            B = normalise_word(B_raw, model)
            C = normalise_word(C_raw, model)
            expected = normalise_word(expected_raw, model)
            #skip when any word is missing
            if None in (A, B, C, expected):
                ws.append([
                    A_raw, B_raw, C_raw,
                    "Not in Vocabulary", expected_raw,
                    0.0, None, False, False, False,
                    None, None, None, None, None, None,
                    None, None, None
                ])
                continue
            
            try:
                #calculate vector analogy (B - A + C)
                result = model.most_similar(
                    positive=[B, C],
                    negative=[A],
                    topn=10 #find top 10 words closest to this vector
                )

                #get predicted words and find rank of expected word
                predicted_words = [w for w, _ in result]
                expected_rank = predicted_words.index(expected) + 1 if expected in predicted_words else None

                #calucalte accuracy metrics
                correct_top1 = (expected_rank == 1)
                correct_top5 = (expected_rank is not None and expected_rank <= 5)
                correct_top10 = (expected_rank is not None and expected_rank <= 10)

                #Get top prediction and the similarity score
                top_word, top_score = result[0]
                #Calculate gender bias metrics for predicted word
                gender_scores = compute_gender_scores(model, top_word)
                #Wirte results to excel sheet
                ws.append([
                    A, B, C,
                    top_word, expected,
                    round(top_score, 3),
                    expected_rank,
                    correct_top1,
                    correct_top5,
                    correct_top10,
                    *gender_scores
                ])

            except KeyError:
                ws.append([
                    A, B, C,
                    "Not in Vocabulary", expected,
                    0.0, None, False, False, False,
                    None, None, None, None, None, None,
                    None, None, None
                ])

            #Progress bar for tracking
            percent = (idx / total_lines) * 100
            print(f"\r{model_name}: {idx}/{total_lines} ({percent:.1f}%)", end="")

            #Save peridocally for data loss prevention
            if idx % save_every == 0:
                wb.save(excel_file)

    except KeyboardInterrupt:
        print("\nInterrupted by user.")
    #Save
    wb.save(excel_file)
    print(f"\nFinished {model_name}.")

#Calculate summary statistics for the models performance
def compute_summary(model_name):

    wb = load_workbook(excel_file)

    sheet_name = f"Analogies_{model_name}"
    summary_sheet_name = f"Summary_{model_name}"

    if sheet_name not in wb.sheetnames:
        return

    #read data rows from sheet
    ws = wb[sheet_name]
    data = list(ws.iter_rows(min_row=2, values_only=True))

    total = 0
    top1 = top5 = top10 = 0
    bias_values = []

    #Processing each row
    for row in data:
        if row[0] is None:
            continue

        total += 1

        if row[7]:  #correct top1
            top1 += 1
        if row[8]:  #correct top5
            top5 += 1
        if row[9]:  #correct top10
            top10 += 1
        #Get bias scores
        if row[18] is not None: #Gender bias score column
            bias_values.append(row[18])

    if total == 0:
        return
    #Calculate the bias statistics
    mean_bias = round(statistics.mean(bias_values), 4) if bias_values else 0
    std_bias = round(statistics.stdev(bias_values), 4) if len(bias_values) > 1 else 0

    #Count words leaning toward each gender
    male_leaning = sum(1 for b in bias_values if b > 0)
    female_leaning = sum(1 for b in bias_values if b < 0)

    #Count strong bias cases, values above threshold
    strong_threshold = 0.1 #medium effect size (typically word similarities fall between -0.5 and 0.8)
    strong_bias = sum(1 for b in bias_values if abs(b) >= strong_threshold)

    #Delete sheet if already exists
    if summary_sheet_name in wb.sheetnames:
        del wb[summary_sheet_name]
    #Create new summary sheet and write statistics
    summary_ws = wb.create_sheet(summary_sheet_name)

    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Total Analogies", total])
    summary_ws.append(["Top1 Accuracy", round(top1 / total, 4)])
    summary_ws.append(["Top5 Accuracy", round(top5 / total, 4)])
    summary_ws.append(["Top10 Accuracy", round(top10 / total, 4)])
    summary_ws.append(["Mean Gender Bias Score", mean_bias])
    summary_ws.append(["Std Dev Gender Bias", std_bias])
    summary_ws.append(["Male-Leaning %", round(male_leaning / len(bias_values), 4) if bias_values else 0])
    summary_ws.append(["Female-Leaning %", round(female_leaning / len(bias_values), 4) if bias_values else 0])
    summary_ws.append([
        f"Strong Bias (|bias| >= {strong_threshold}) %",
        round(strong_bias / len(bias_values), 4) if bias_values else 0
    ])

    wb.save(excel_file)
    print(f"Summary written for {model_name}")

#Loop until user exits program
while True:
        #select or create
        excel_file = choose_excel(excel_dir)

        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Placeholder" #Placeholder file to avoid no sheet error
            wb.save(excel_file)
            print(f"\nCreated new Excel file: {excel_file}")
        else:
            print(f"\nUsing existing Excel file: {excel_file}")

        #Choose the text file with analogies
        txt_file = choose_txt(text_dir)

        if not os.path.exists(txt_file):
            print("Text file not found.")
            continue

        print(f"\nUsing text file: {txt_file}")

        #Process text file for each model
        for model_name, model in models.items():
            run_analogy(txt_file, model, model_name)
            compute_summary(model_name)

        print("\nFinished processing this file.")

        again = input("\nRun another text file? (y/n): ").strip().lower()

        if again != "y":
            print("Exiting.")
            break