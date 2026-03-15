import gensim.downloader as api
from gensim.models import KeyedVectors
from openpyxl import Workbook, load_workbook
import os
import numpy as np

#Setting up directory paths relative to script
current_dir = os.path.dirname(os.path.abspath(__file__))
text_dir = os.path.join(current_dir, "text files")
excel_dir = os.path.join(current_dir, "all excel files")

#load pretrained embedding models
models = {}
#Google News W2V model
print("Word2Vec loading")
models["word2vec"] = api.load("word2vec-google-news-300")
print("Word2Vec loaded.")

#load glove model in Keyed Vectors format
print("GloVe loading")
glove_fast_path = os.path.join(current_dir, "glove_fast.kv")
models["glove"] = KeyedVectors.load(glove_fast_path)
print("GloVe loaded.")

#prompt user to select an existing excel file or create a new one
def choose_excel(directory):
    excel_files = [f for f in os.listdir(directory) if f.endswith(".xlsx")]

    print("\nExcel file options:")
    if excel_files:
        for i, f in enumerate(excel_files, start=1):
            print(f"  {i}. Use existing file: {f}")

    print(f"  {len(excel_files) + 1}. Create a new Excel file")

    while True:
        choice = input("\nSelect an option: ").strip()

        if choice.isdigit():
            choice = int(choice)

            if 1 <= choice <= len(excel_files):
                return os.path.join(directory, excel_files[choice - 1]) #path to chosen file

            if choice == len(excel_files) + 1:
                name = input("Enter new Excel filename (without .xlsx): ").strip()  #path to new
                return os.path.join(directory, f"{name}.xlsx")

        print("Invalid choice. Please try again.")

#Loop until user exits program
while True:
    #Select or crrate new Excel file
    excel_file = choose_excel(excel_dir)
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Placeholder" #Placeholder sheet
        wb.save(excel_file)
        print(f"\nCreated new Excel file: {excel_file}")
    else:
        print(f"\nUsing existing Excel file: {excel_file}")

    #Show available text files and prompt user to choose
    def choose_txt(directory):
        txt_files = [f for f in os.listdir(directory) if f.endswith(".txt")]

        if not txt_files:
            print("No .txt files found.")
            exit()

        print("\nText file options:")
        for i, f in enumerate(txt_files, start=1):
            print(f"  {i}. {f}")

        while True:
            choice = input("\nSelect a Text file: ").strip()
            if choice.isdigit():
                choice = int(choice)
                if 1 <= choice <= len(txt_files):
                    return os.path.join(directory, txt_files[choice - 1])

            print("Invalid choice. Please try again.")

    txt_file = choose_txt(text_dir)
    print(f"\nUsing file: {txt_file}")

    #Normalise the word to find match in model
    def normalise_word(word, model):
        for form in [word, word.lower(), word.capitalize(), word.upper()]:
            if form in model:
                return form
        return None

    #Create a vector representation for multiword phrases by averaging the vectors of individual words
    def phrase_vector(model, phrase):
        #Convert underscores to spaces
        phrase = phrase.replace("_", " ")

        words = phrase.split()
        vectors = []

        #Get vector for each word that exsists
        for w in words:
            w_norm = normalise_word(w, model)
            if w_norm and w_norm in model:
                vectors.append(model[w_norm])

        if not vectors:
            return None
        #Return the average
        return np.mean(vectors, axis=0)

    #Calculate the cosine similarity between a vector and word
    def cosine_similarity_vec_word(model, vec1, w2):
        w2_norm = normalise_word(w2, model)

        if vec1 is None or w2_norm not in model:
            return None

        vec2 = model[w2_norm]
        #Cosine similarity calculation
        return round(
            np.dot(vec1, vec2) /
            (np.linalg.norm(vec1) * np.linalg.norm(vec2)),
            3
        )

    #Get or create a sheet for specific model
    def get_create_sheet(wb, model_name):

        sheet_name = f"Bias_{model_name}"

        if sheet_name not in wb.sheetnames:
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet":
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(sheet_name)
            #Write column headers
            ws.append([
                "Word", #word being analysed
                "Sim_to_Male",  #cosine simalirty to "male"
                "Sim_to_Female",
                "Sim_to_Man",
                "Sim_to_Woman",
                "Sim_to_Boy",
                "Sim_to_Girl",
                "Avg_Male_Group",   #avg similarity across set of male words
                "Avg_Female_Group", #avg similarity across set of female words
                "Gender_Bias_Score" #avg male - avg female
            ])

            wb.save(excel_file)

        return wb[sheet_name]

    #Calculate summary statistics for bias analysis of a model 
    def compute_summary(model_name):

        wb = load_workbook(excel_file)

        sheet_name = f"Bias_{model_name}"
        summary_sheet_name = f"Summary_{model_name}"

        if sheet_name not in wb.sheetnames:
            return

        #Read data from sheet
        ws = wb[sheet_name]
        data = list(ws.iter_rows(min_row=2, values_only=True))

        total = 0
        bias_values = []

        #Prcoessing bias scores from each row
        for row in data:
            word = row[0]
            bias_score = row[9] #Gender bias score column

            if word is None:
                continue

            total += 1

            if bias_score is not None:
                bias_values.append(bias_score)

        if total == 0:
            return
        #Calculate the bias statistics
        mean_bias = round(np.mean(bias_values), 4) if bias_values else 0
        std_bias = round(np.std(bias_values), 4) if len(bias_values) > 1 else 0

        #Count words leaning toward each gender
        male_leaning = sum(1 for b in bias_values if b > 0)
        female_leaning = sum(1 for b in bias_values if b < 0)

         #Count strong bias cases, values above threshold
        strong_threshold = 0.1
        strong_bias = sum(1 for b in bias_values if abs(b) >= strong_threshold) #medium effect size (typically word similarities fall between -0.5 and 0.8)

        #Delete sheet if already exists
        if summary_sheet_name in wb.sheetnames:
            del wb[summary_sheet_name]
        #Create new summary sheet and write statistics
        summary_ws = wb.create_sheet(summary_sheet_name)

        summary_ws.append(["Metric", "Value"])
        summary_ws.append(["Total Words", total])
        summary_ws.append(["Mean Gender Bias Score", mean_bias])
        summary_ws.append(["Std Dev Gender Bias", std_bias])
        summary_ws.append(["Male-Leaning %", round(male_leaning / len(bias_values), 4) if bias_values else 0])
        summary_ws.append(["Female-Leaning %", round(female_leaning / len(bias_values), 4) if bias_values else 0])
        summary_ws.append([
            f"Strong Bias (|bias| >= {strong_threshold}) %",
            round(strong_bias / len(bias_values), 4) if bias_values else 0
        ])

        wb.save(excel_file)
        print(f"Summary written for {model_name}")

    #Read all lines from the text file
    with open(txt_file, "r", encoding="utf-8") as f:
        words_list = [line.strip() for line in f if line.strip()]

    #Save excel file every 10 words to prevent data loss
    save_every = 10

    for model_name, model in models.items():

        print(f"\nProcessing {model_name}...")

        #Load workbook and get or create sheet for the model
        wb = load_workbook(excel_file)
        ws = get_create_sheet(wb, model_name)

        #determine how many lines already processed
        processed_words = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                processed_words.add(row[0])

        try:
            for idx, word in enumerate(words_list, start=1):
                #Skip if already processed
                if word in processed_words:
                    continue
                
                #Get vector representation
                vec_word = phrase_vector(model, word)
                #Out of vocabulary
                if vec_word is None:
                    ws.append([word] + [None]*9)
                    continue
                #Calculate similarities to gendered terms
                sim_male   = cosine_similarity_vec_word(model, vec_word, "male")
                sim_female = cosine_similarity_vec_word(model, vec_word, "female")
                sim_man    = cosine_similarity_vec_word(model, vec_word, "man")
                sim_woman  = cosine_similarity_vec_word(model, vec_word, "woman")
                sim_boy    = cosine_similarity_vec_word(model, vec_word, "boy")
                sim_girl   = cosine_similarity_vec_word(model, vec_word, "girl")

                #Group similarities by gender
                male_group = [x for x in [sim_male, sim_man, sim_boy] if x is not None]
                female_group = [x for x in [sim_female, sim_woman, sim_girl] if x is not None]

                #Calculate average similarity for each gender group
                avg_male = round(sum(male_group) / len(male_group), 3) if male_group else None
                avg_female = round(sum(female_group) / len(female_group), 3) if female_group else None

                #Calculate gender bias score
                gender_bias_score = (
                    round(avg_male - avg_female, 3)
                    if avg_male is not None and avg_female is not None
                    else None
                )
                #Write results to sheet
                ws.append([
                    word,
                    sim_male,
                    sim_female,
                    sim_man,
                    sim_woman,
                    sim_boy,
                    sim_girl,
                    avg_male,
                    avg_female,
                    gender_bias_score
                ])

                #Save periodically
                if idx % save_every == 0:
                    wb.save(excel_file)

        except KeyboardInterrupt:
            print("\nInterrupted by user.")
        #Save and generate summaries
        wb.save(excel_file)
        compute_summary(model_name)

    print("\nFinished processing this file.")

    again = input("\nRun another text file? (y/n): ").strip().lower()

    if again != "y":
        print("Exiting.")
        break